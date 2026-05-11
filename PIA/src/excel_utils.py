import json
import os
from openpyxl import Workbook

def generar_excel_actividad(json_path, excel_path):
    if not os.path.exists(json_path):
        print("Error: JSON no encontrado.")
        return

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "datos limpios"
    ws1.append(["Año", "Autor", "Tipo"]) 
    
    datos_por_año = data.get("por_año", {})
    for año, contenido in datos_por_año.items():
        autores = contenido.get("autores", [])
        tipos = contenido.get("tipos", [])
        max_len = max(len(autores), len(tipos))
        for i in range(max_len):
            autor = autores[i] if i < len(autores) else ""
            tipo = tipos[i] if i < len(tipos) else ""
            ws1.append([año, autor, tipo])

    ws2 = wb.create_sheet("estadísticas")
    ws2.append(["Año", "Cantidad Total de Publicaciones"])
    resumen = data.get("publicaciones_por_año", [])
    for item in resumen:
        ws2.append([item.get("Año"), item.get("Cantidad")])

    ws3 = wb.create_sheet("tabla de frecuencia")
    ws3.append(["Autor", "Frecuencia (Apariciones)"])
    
    todos_autores = data.get("total", {}).get("autores", [])
    frecuencias = {}
    for a in todos_autores:
        frecuencias[a] = frecuencias.get(a, 0) + 1
    
    for autor, count in frecuencias.items():
        ws3.append([autor, count])

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb.save(excel_path)
    print(f"Excel creado exitosamente en: {excel_path}")

