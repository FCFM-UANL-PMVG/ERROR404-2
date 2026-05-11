import json
import matplotlib.pyplot as plt
import os
from collections import Counter
import statistics as stats
from openpyxl import Workbook

def generar_excel_actividad(json_path, excel_path):
    if not os.path.exists(json_path):
        print("Error: JSON no encontrado.")
        return

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "datos limpios"
    ws1.append(["Año", "Autor", "Tipo"]) 
    
    datos_por_año = data.get("por_año", {})
    for año, contenido in datos_por_año.items():
        autores = contenido.get("autores", [])
        tipos = contenido.get("tipos", [])
        max_len = max(len(autores), len(tipos))
        for i in range(max_len):
            autor = autores[i] if i < len(autores) else ""
            tipo = tipos[i] if i < len(tipos) else ""
            ws1.append([año, autor, tipo])

    ws2 = wb.create_sheet("estadísticas")
    ws2.append(["Año", "Cantidad Total de Publicaciones"])
    resumen = data.get("publicaciones_por_año", [])
    for item in resumen:
        ws2.append([item.get("Año"), item.get("Cantidad")])

    ws3 = wb.create_sheet("tabla de frecuencia")
    ws3.append(["Autor", "Frecuencia (Apariciones)"])
    
    todos_autores = data.get("total", {}).get("autores", [])
    frecuencias = {}
    for a in todos_autores:
        frecuencias[a] = frecuencias.get(a, 0) + 1
    
    for autor, count in frecuencias.items():
        ws3.append([autor, count])

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb.save(excel_path)
    print(f"Excel creado exitosamente en: {excel_path}")


def graficas():
    fechas_investigadas = [2001, 2002, 2003, 2004, 2005, 2006, 2007, 2008]
    publicaciones = articulos["publicaciones_por_año"]

    try:
        if len(publicaciones) == len(fechas_investigadas):
            suma_publi = 0
            for p in publicaciones:
                suma_publi += p
            porcentaje = []
            for p in publicaciones:
                porcen = p / suma_publi
                porcentaje.append(porcen)

            plt.figure()
            plt.title("Publicaciones por Año")
            plt.pie(porcentaje, labels=fechas_investigadas, autopct='%1.1f%%')
            plt.savefig("figures/pie_porcentaje_global.png")
            
        else:
            print("Las listas de años y porcentajes no tiene la misma longitud")

        lista_autor = articulos["total"]["autores"]
        autores = Counter(lista_autor)
        autores_mayor = autores.most_common(3)
        nombre_autores = []
        cantidad_autores = []
        for par in autores_mayor:
            nombress = par[0]
            nombre_autores.append(nombress)
            cantidadd = par[1]
            cantidad_autores.append(cantidadd)
            
        plt.figure()
        plt.title("Top 3 autores con más publicaciones")
        plt.xlabel("Autor")
        plt.ylabel("Publicaciones")
        plt.bar(nombre_autores, cantidad_autores, color="blue")
        plt.savefig("figures/bar_top3.png")


        formatos_año = {
            "PDF": [],
            "HTML": [],
            "Libros": [],
            "Otros": []
        }

        for fecha in fechas_investigadas:
            fecha_str = str(fecha)
            formatos = articulos["por_año"].get(fecha_str).get("tipos", [])
            format_count = Counter(formatos)
            formatos_año["PDF"].append(format_count.get("Pdf", 0))
            formatos_año["HTML"].append(format_count.get("Html", 0))
            formatos_año["Libros"].append(format_count.get("Books", 0))
            formatos_año["Otros"].append(format_count.get("Otros", 0))
        
        plt.figure()
        plt.title("Formatos utilizados")
        plt.xlabel("Año")
        plt.ylabel("Cantidad")
        plt.plot(fechas_investigadas, formatos_año["PDF"], label="PDF")
        plt.plot(fechas_investigadas, formatos_año["HTML"], label="HTML")
        plt.plot(fechas_investigadas, formatos_año["Libros"], label="Libros")
        plt.plot(fechas_investigadas, formatos_año["Otros"], label="Otros")
        plt.savefig("figures/plot_formatos.png")


        plt.figure()
        plt.title("Publicaciones por año")
        plt.xlabel("Años")
        plt.ylabel("Cantidad")
        plt.bar(publicaciones, fechas_investigadas, color="pink")
        
        
        plt.show()
        
    except FileNotFoundError:
        print("No se encontró el archivo, intente correr el script principal primero.")

graficando = graficas()
